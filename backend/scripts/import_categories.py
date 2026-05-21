"""Excel 三级分类导入 CLI 脚本(对齐 docs/商品三级分类-PRD-v1.0.md §4)。

用法
----
    # 默认从项目根 data/ 下取 mtime 最新的 .xlsx
    python scripts/import_categories.py

    # 显式指定文件(路径必须在 data/ 下)
    python scripts/import_categories.py --file ../data/三局产品三级分类（整合合同分类）20260516.xlsx

    # 只看差异不写库
    python scripts/import_categories.py --dry-run

    # 把 DB 里有但本次 Excel 没有的分类置 is_active=false
    python scripts/import_categories.py --deactivate-missing

设计要点
--------
- code 永久不变契约(PRD §3.4):按 (name_zh, parent_code) 匹配现有节点,沿用 code
- L1 列 Excel 有"省略写法"(只在该一级第一行有值,后续行 None),需向下填充
- L3 单元格可多值,用 、,，;；\\n\\r 拆分
- 路径强制在 data/ 下,fail-fast
- 永不物理删,删除走 is_active=false

⚠️ 本脚本**不在应用启动时自动跑**,只能本地人工执行。
"""
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

# 让脚本能 import app.*
_BACKEND_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_BACKEND_ROOT))

from app.core.config import settings  # noqa: E402
from app.db.base import _utcnow  # noqa: E402
from app.db.models import Category  # noqa: E402
from app.db.url import prepare_sync_url  # noqa: E402

PROJECT_ROOT = _BACKEND_ROOT.parent
DATA_DIR = PROJECT_ROOT / "data"

# L3 单元格多值分隔符(PRD §3.3 D2 决策已锁定)
L3_SPLIT_PATTERN = re.compile(r"[、,，;；\n\r]+")

# 表头模糊匹配关键词
HEADER_L1_KEYS = ("一级",)
HEADER_L2_KEYS = ("二级",)
HEADER_L3_KEYS = ("三级",)


# ---------------- 数据结构 ----------------


@dataclass
class ExcelL2:
    name_zh: str
    l3_names: list[str] = field(default_factory=list)


@dataclass
class ExcelL1:
    name_zh: str
    l2_nodes: list[ExcelL2] = field(default_factory=list)


@dataclass
class ExcelTree:
    l1_nodes: list[ExcelL1] = field(default_factory=list)


@dataclass
class ImportStats:
    inserted: int = 0
    updated: int = 0
    # DB 有但 Excel 无,默认保留;加 --deactivate-missing 才停用
    kept: int = 0
    deactivated: int = 0

    inserted_codes: list[str] = field(default_factory=list)
    deactivated_codes: list[str] = field(default_factory=list)


# ---------------- 文件定位与校验 ----------------


def find_latest_xlsx_in_data_dir() -> Path:
    """扫 data/ 下 *.xlsx,取 mtime 最新的一份;歧义/不存在时 fail-fast。

    跳过以 `.` 或 `~$` 开头的隐藏/lock 文件(macOS / Office 临时文件)。
    """
    if not DATA_DIR.is_dir():
        sys.exit(f"[ERROR] 数据目录不存在: {DATA_DIR}")
    candidates = sorted(
        (
            p
            for p in DATA_DIR.glob("*.xlsx")
            if not p.name.startswith((".", "~$"))
        ),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        sys.exit(f"[ERROR] {DATA_DIR} 下没有可用的 .xlsx 文件")
    if len(candidates) > 1:
        print("[INFO] data/ 下存在多份 xlsx,按 mtime 取最新:")
        for p in candidates:
            mark = "  ← 选中" if p == candidates[0] else ""
            print(f"  {p.name}{mark}")
    return candidates[0]


def validate_xlsx_path(path: Path) -> Path:
    """校验路径在 data/ 下且为 .xlsx,否则 fail-fast。返回 resolve 后的绝对路径。"""
    abs_path = path.resolve()
    if not abs_path.exists():
        sys.exit(f"[ERROR] 文件不存在: {abs_path}")
    try:
        abs_path.relative_to(DATA_DIR.resolve())
    except ValueError:
        sys.exit(
            "[ERROR] Excel 路径必须在 data/ 目录下\n"
            f"        允许范围: {DATA_DIR}\n"
            f"        实际路径: {abs_path}"
        )
    if abs_path.suffix.lower() != ".xlsx":
        sys.exit(f"[ERROR] 仅支持 .xlsx 格式: {abs_path}")
    return abs_path


# ---------------- Excel 解析 ----------------


def _match_header(headers: tuple, keys: tuple[str, ...]) -> int | None:
    """模糊匹配表头,返回列序号(0-based);找不到返回 None。"""
    for idx, cell in enumerate(headers):
        if cell is None:
            continue
        s = str(cell).strip().lower()
        if any(k in s for k in keys):
            return idx
    return None


def parse_xlsx(path: Path) -> ExcelTree:
    """读 Excel,模糊匹配表头,L1 向下填充,L3 多值拆分,按出现顺序去重。"""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        sys.exit(f"[ERROR] Excel 无活动 sheet: {path}")

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if headers is None:
        sys.exit(f"[ERROR] Excel 为空: {path}")

    l1_col = _match_header(headers, HEADER_L1_KEYS)
    l2_col = _match_header(headers, HEADER_L2_KEYS)
    l3_col = _match_header(headers, HEADER_L3_KEYS)
    if l1_col is None or l2_col is None or l3_col is None:
        sys.exit(
            "[ERROR] 表头未匹配上 一级/二级/三级 分类\n"
            f"        实际表头: {headers}"
        )

    tree = ExcelTree()
    l1_index: dict[str, ExcelL1] = {}
    l2_index: dict[tuple[str, str], ExcelL2] = {}
    seen_l3: set[tuple[str, str, str]] = set()

    last_l1: str | None = None
    for row_num, row in enumerate(rows_iter, start=2):
        l1_raw = row[l1_col]
        l2_raw = row[l2_col]
        l3_raw = row[l3_col]

        # 整行空跳过
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in (l1_raw, l2_raw, l3_raw)):
            continue

        # L1 向下填充
        if l1_raw is not None and str(l1_raw).strip():
            last_l1 = str(l1_raw).strip()
        if not last_l1:
            sys.exit(f"[ERROR] 第 {row_num} 行 一级分类 缺失且无上文可填: {row}")
        l1_name = last_l1

        if l2_raw is None or not str(l2_raw).strip():
            sys.exit(f"[ERROR] 第 {row_num} 行 二级分类 缺失: {row}")
        l2_name = str(l2_raw).strip()

        # L1 节点
        if l1_name not in l1_index:
            node = ExcelL1(name_zh=l1_name)
            l1_index[l1_name] = node
            tree.l1_nodes.append(node)
        l1_node = l1_index[l1_name]

        # L2 节点
        l2_key = (l1_name, l2_name)
        if l2_key not in l2_index:
            l2_node = ExcelL2(name_zh=l2_name)
            l2_index[l2_key] = l2_node
            l1_node.l2_nodes.append(l2_node)
        l2_node = l2_index[l2_key]

        # L3 多值拆分(空 cell 也允许:表示该 L2 暂无 L3)
        if l3_raw is None:
            continue
        for raw in L3_SPLIT_PATTERN.split(str(l3_raw)):
            name = raw.strip()
            if not name:
                continue
            triple = (l1_name, l2_name, name)
            if triple in seen_l3:
                continue
            seen_l3.add(triple)
            l2_node.l3_names.append(name)

    wb.close()
    return tree


# ---------------- code 生成 ----------------


def _split_seq(code: str) -> int:
    """从 code 末尾段取整数序号:'01'→1, '01.005'→5, '01.005.012'→12。"""
    return int(code.split(".")[-1])


def _next_seq(used: set[int]) -> int:
    seq = 1
    while seq in used:
        seq += 1
    return seq


def _make_code(parent: str | None, seq: int, level: int) -> str:
    if level == 1:
        return f"{seq:02d}"
    assert parent is not None
    return f"{parent}.{seq:03d}"


# ---------------- 核心导入算法(PRD §4.5) ----------------


def _upsert_existing(
    cat: Category, name_zh: str, sort_order: int, dry_run: bool
) -> bool:
    """更新现有 category,有变化返回 True。dry-run 时只判断,不写。"""
    changed = (
        cat.name_zh != name_zh
        or cat.sort_order != sort_order
        or not cat.is_active
    )
    if changed and not dry_run:
        cat.name_zh = name_zh
        cat.sort_order = sort_order
        cat.is_active = True
        cat.updated_at = _utcnow()
    return changed


def _insert_new(
    db: Session,
    *,
    code: str,
    name_zh: str,
    level: int,
    parent_code: str | None,
    sort_order: int,
    dry_run: bool,
) -> None:
    if dry_run:
        return
    now = _utcnow()
    db.add(
        Category(
            code=code,
            name_zh=name_zh,
            level=level,
            parent_code=parent_code,
            sort_order=sort_order,
            is_active=True,
            created_at=now,
            updated_at=now,
        )
    )


def import_from_xlsx(
    db: Session,
    tree: ExcelTree,
    dry_run: bool = False,
    deactivate_missing: bool = False,
) -> ImportStats:
    """核心算法:沿用已有 code,新增节点取空号(PRD §4.5)。"""
    stats = ImportStats()

    # 加载现有所有 categories,按自然键索引
    existing_by_natural: dict[tuple[str, str | None], Category] = {}
    existing_by_code: dict[str, Category] = {}
    used_seq_by_parent: dict[str | None, set[int]] = {}

    for c in db.execute(select(Category)).scalars().all():
        existing_by_natural[(c.name_zh, c.parent_code)] = c
        existing_by_code[c.code] = c
        used_seq_by_parent.setdefault(c.parent_code, set()).add(_split_seq(c.code))

    excel_codes: set[str] = set()

    for sort_l1, l1 in enumerate(tree.l1_nodes):
        l1_existing = existing_by_natural.get((l1.name_zh, None))
        if l1_existing:
            l1_code = l1_existing.code
            if _upsert_existing(l1_existing, l1.name_zh, sort_l1, dry_run):
                stats.updated += 1
        else:
            used = used_seq_by_parent.setdefault(None, set())
            seq = _next_seq(used)
            used.add(seq)
            l1_code = _make_code(None, seq, 1)
            _insert_new(
                db,
                code=l1_code,
                name_zh=l1.name_zh,
                level=1,
                parent_code=None,
                sort_order=sort_l1,
                dry_run=dry_run,
            )
            stats.inserted += 1
            stats.inserted_codes.append(l1_code)
        excel_codes.add(l1_code)

        for sort_l2, l2 in enumerate(l1.l2_nodes):
            l2_existing = existing_by_natural.get((l2.name_zh, l1_code))
            if l2_existing:
                l2_code = l2_existing.code
                if _upsert_existing(l2_existing, l2.name_zh, sort_l2, dry_run):
                    stats.updated += 1
            else:
                used = used_seq_by_parent.setdefault(l1_code, set())
                seq = _next_seq(used)
                used.add(seq)
                l2_code = _make_code(l1_code, seq, 2)
                _insert_new(
                    db,
                    code=l2_code,
                    name_zh=l2.name_zh,
                    level=2,
                    parent_code=l1_code,
                    sort_order=sort_l2,
                    dry_run=dry_run,
                )
                stats.inserted += 1
                stats.inserted_codes.append(l2_code)
            excel_codes.add(l2_code)

            for sort_l3, l3_name in enumerate(l2.l3_names):
                l3_existing = existing_by_natural.get((l3_name, l2_code))
                if l3_existing:
                    l3_code = l3_existing.code
                    if _upsert_existing(l3_existing, l3_name, sort_l3, dry_run):
                        stats.updated += 1
                else:
                    used = used_seq_by_parent.setdefault(l2_code, set())
                    seq = _next_seq(used)
                    used.add(seq)
                    l3_code = _make_code(l2_code, seq, 3)
                    _insert_new(
                        db,
                        code=l3_code,
                        name_zh=l3_name,
                        level=3,
                        parent_code=l2_code,
                        sort_order=sort_l3,
                        dry_run=dry_run,
                    )
                    stats.inserted += 1
                    stats.inserted_codes.append(l3_code)
                excel_codes.add(l3_code)

    # 处理 DB 有但 Excel 没有的节点
    for code, cat in existing_by_code.items():
        if code in excel_codes:
            continue
        if deactivate_missing:
            stats.deactivated += 1
            stats.deactivated_codes.append(code)
            if not dry_run and cat.is_active:
                cat.is_active = False
                cat.updated_at = _utcnow()
        else:
            stats.kept += 1

    return stats


# ---------------- CLI ----------------


def _print_stats(stats: ImportStats, dry_run: bool) -> None:
    title = "DRY RUN 差异统计" if dry_run else "导入结果"
    print(f"--- {title} ---")
    print(f"新增   : {stats.inserted}")
    print(f"更新   : {stats.updated}")
    print(
        f"保留不动: {stats.kept}    (DB 有但 Excel 无;加 --deactivate-missing 将停用)"
    )
    print(f"将停用 : {stats.deactivated}")
    if stats.inserted_codes:
        n = min(10, len(stats.inserted_codes))
        more = (
            f"  ... (+{len(stats.inserted_codes) - n})"
            if len(stats.inserted_codes) > n
            else ""
        )
        print(f"  新增 codes 样例: {stats.inserted_codes[:n]}{more}")
    if stats.deactivated_codes:
        n = min(10, len(stats.deactivated_codes))
        more = (
            f"  ... (+{len(stats.deactivated_codes) - n})"
            if len(stats.deactivated_codes) > n
            else ""
        )
        print(f"  停用 codes 样例: {stats.deactivated_codes[:n]}{more}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="商品三级分类 Excel 导入(对齐 PRD §4)"
    )
    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="Excel 文件路径(必须在 data/ 下);省略则取 data/ 下 mtime 最新的 .xlsx",
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="只打印差异统计,不写库"
    )
    parser.add_argument(
        "--deactivate-missing",
        action="store_true",
        help="把 DB 里有但 Excel 没有的分类置 is_active=false",
    )
    args = parser.parse_args()

    if args.file is None:
        xlsx_path = find_latest_xlsx_in_data_dir()
    else:
        xlsx_path = args.file
    xlsx_path = validate_xlsx_path(xlsx_path)

    print(f"[INFO] Excel: {xlsx_path}")
    print(
        f"[INFO] dry-run={args.dry_run}, deactivate-missing={args.deactivate_missing}"
    )

    tree = parse_xlsx(xlsx_path)
    n_l1 = len(tree.l1_nodes)
    n_l2 = sum(len(n.l2_nodes) for n in tree.l1_nodes)
    n_l3 = sum(len(l2.l3_names) for l1 in tree.l1_nodes for l2 in l1.l2_nodes)
    print(
        f"[INFO] Excel 解析: L1={n_l1}, L2={n_l2}, L3={n_l3}, 总计={n_l1 + n_l2 + n_l3}"
    )

    sync_url = prepare_sync_url(settings.DATABASE_URL)
    engine = create_engine(sync_url)
    with Session(engine) as db:
        stats = import_from_xlsx(
            db,
            tree,
            dry_run=args.dry_run,
            deactivate_missing=args.deactivate_missing,
        )
        if args.dry_run:
            db.rollback()
            print("[DRY RUN] 未写库,事务已 rollback")
        else:
            db.commit()
            print("[OK] 已 commit")

    _print_stats(stats, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
